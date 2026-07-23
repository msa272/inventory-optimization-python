import math
import os

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# -------------------------------------------------
# DISPLAY SETTINGS
# -------------------------------------------------

pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)


# -------------------------------------------------
# CREATE CHARTS FOLDER
# -------------------------------------------------

os.makedirs("charts", exist_ok=True)


# -------------------------------------------------
# READ DATA
# -------------------------------------------------

inventory = pd.read_csv("inventory_data.csv")


# -------------------------------------------------
# INVENTORY VALUE
# -------------------------------------------------

inventory["Inventory_Value"] = (
    inventory["Annual_Demand"]
    * inventory["Unit_Cost"]
)

total_value = inventory["Inventory_Value"].sum()


# -------------------------------------------------
# ABC ANALYSIS
# -------------------------------------------------

inventory = inventory.sort_values(
    by="Inventory_Value",
    ascending=False
)

inventory["Cumulative_Value"] = (
    inventory["Inventory_Value"].cumsum()
)

inventory["Cumulative_%"] = (
    inventory["Cumulative_Value"]
    / inventory["Inventory_Value"].sum()
) * 100


def classify(percent):
    if percent <= 80:
        return "A"
    elif percent <= 95:
        return "B"
    else:
        return "C"


inventory["ABC_Class"] = (
    inventory["Cumulative_%"].apply(classify)
)


# -------------------------------------------------
# ECONOMIC ORDER QUANTITY
# -------------------------------------------------

inventory["EOQ"] = inventory.apply(
    lambda row: math.sqrt(
        (
            2
            * row["Annual_Demand"]
            * row["Ordering_Cost"]
        )
        / row["Holding_Cost"]
    ),
    axis=1
)

inventory["EOQ"] = inventory["EOQ"].round(2)


# -------------------------------------------------
# SAFETY STOCK
# -------------------------------------------------

inventory["Safety_Stock"] = (
    inventory["Service_Level_Z"]
    * inventory["Daily_Demand_StdDev"]
    * inventory["Lead_Time"].apply(math.sqrt)
)

inventory["Safety_Stock"] = (
    inventory["Safety_Stock"].round(2)
)


# -------------------------------------------------
# REORDER POINT
# -------------------------------------------------

inventory["Daily_Demand"] = (
    inventory["Annual_Demand"] / 365
)

inventory["Reorder_Point"] = (
    inventory["Daily_Demand"]
    * inventory["Lead_Time"]
) + inventory["Safety_Stock"]

inventory["Daily_Demand"] = (
    inventory["Daily_Demand"].round(2)
)

inventory["Reorder_Point"] = (
    inventory["Reorder_Point"].round(2)
)


# -------------------------------------------------
# FINAL TERMINAL OUTPUT
# -------------------------------------------------

print("\n" + "=" * 120)

print(
    "ABC ANALYSIS, EOQ, SAFETY STOCK "
    "& REORDER POINT"
)

print("=" * 120)

print(
    inventory[
        [
            "Product",
            "Inventory_Value",
            "ABC_Class",
            "EOQ",
            "Safety_Stock",
            "Reorder_Point"
        ]
    ]
)

print("\n" + "=" * 60)
print(f"Total Inventory Value: ${total_value:,.2f}")
print("=" * 60)


# -------------------------------------------------
# CHART 1: INVENTORY VALUE
# -------------------------------------------------

plt.figure(figsize=(10, 6))

plt.bar(
    inventory["Product"],
    inventory["Inventory_Value"]
)

plt.title("Inventory Value by Product")
plt.xlabel("Product")
plt.ylabel("Inventory Value")

plt.xticks(
    rotation=45,
    ha="right"
)

plt.tight_layout()

plt.savefig(
    "charts/inventory_value.png",
    dpi=300
)

plt.close()


# -------------------------------------------------
# CHART 2: EOQ
# -------------------------------------------------

plt.figure(figsize=(10, 6))

plt.bar(
    inventory["Product"],
    inventory["EOQ"]
)

plt.title("Economic Order Quantity by Product")
plt.xlabel("Product")
plt.ylabel("EOQ")

plt.xticks(
    rotation=45,
    ha="right"
)

plt.tight_layout()

plt.savefig(
    "charts/eoq_chart.png",
    dpi=300
)

plt.close()


# -------------------------------------------------
# CHART 3: SAFETY STOCK AND REORDER POINT
# -------------------------------------------------

chart_data = inventory.set_index("Product")[
    [
        "Safety_Stock",
        "Reorder_Point"
    ]
]

chart_data.plot(
    kind="bar",
    figsize=(10, 6)
)

plt.title("Safety Stock and Reorder Point")
plt.xlabel("Product")
plt.ylabel("Units")

plt.xticks(
    rotation=45,
    ha="right"
)

plt.tight_layout()

plt.savefig(
    "charts/safety_stock_reorder_point.png",
    dpi=300
)

plt.close()


# -------------------------------------------------
# CHART 4: PARETO CHART
# -------------------------------------------------

fig, ax1 = plt.subplots(figsize=(10, 6))

ax1.bar(
    inventory["Product"],
    inventory["Inventory_Value"]
)

ax1.set_xlabel("Product")
ax1.set_ylabel("Inventory Value")

ax1.tick_params(
    axis="x",
    rotation=45
)

ax2 = ax1.twinx()

ax2.plot(
    inventory["Product"],
    inventory["Cumulative_%"],
    marker="o"
)

ax2.set_ylabel("Cumulative Percentage")

ax2.axhline(
    y=80,
    linestyle="--"
)

ax2.axhline(
    y=95,
    linestyle="--"
)

plt.title("Inventory Value Pareto Analysis")

fig.tight_layout()

plt.savefig(
    "charts/pareto_chart.png",
    dpi=300
)

plt.close()

print("\nCharts created successfully in the 'charts' folder.")


# -------------------------------------------------
# EXPORT FORMATTED EXCEL REPORT
# -------------------------------------------------

excel_file = "inventory_report.xlsx"

inventory.to_excel(
    excel_file,
    index=False
)

workbook = load_workbook(excel_file)
worksheet = workbook.active

worksheet.title = "Inventory Analysis"


# Header formatting
header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font


# Auto-fit columns
for column in worksheet.columns:
    max_length = 0
    column_letter = get_column_letter(
        column[0].column
    )

    for cell in column:
        if cell.value is not None:
            cell_length = len(str(cell.value))

            if cell_length > max_length:
                max_length = cell_length

    worksheet.column_dimensions[
        column_letter
    ].width = max_length + 3


# Freeze header row
worksheet.freeze_panes = "A2"


# Add number formatting
currency_columns = [
    "Unit_Cost",
    "Ordering_Cost",
    "Holding_Cost",
    "Inventory_Value",
    "Cumulative_Value"
]

decimal_columns = [
    "Cumulative_%",
    "EOQ",
    "Safety_Stock",
    "Daily_Demand",
    "Reorder_Point"
]

header_positions = {
    cell.value: cell.column
    for cell in worksheet[1]
}

for column_name in currency_columns:
    if column_name in header_positions:
        column_number = header_positions[column_name]

        for row_number in range(
            2,
            worksheet.max_row + 1
        ):
            worksheet.cell(
                row=row_number,
                column=column_number
            ).number_format = '$#,##0.00'

for column_name in decimal_columns:
    if column_name in header_positions:
        column_number = header_positions[column_name]

        for row_number in range(
            2,
            worksheet.max_row + 1
        ):
            worksheet.cell(
                row=row_number,
                column=column_number
            ).number_format = '0.00'


# Add filter to header row
worksheet.auto_filter.ref = worksheet.dimensions


# Save formatted workbook
workbook.save(excel_file)

print("Professional Excel report created successfully!")